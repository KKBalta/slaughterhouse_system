import logging
import statistics
from collections import defaultdict
from datetime import datetime, time, timedelta
from zoneinfo import ZoneInfo

from django.db.models import Count

logger = logging.getLogger(__name__)


class ReportDataAggregator:
    """Central service for aggregating data from all apps"""

    def __init__(self, start_date, end_date, filters=None):
        self.start_date = start_date
        self.end_date = end_date
        self.filters = filters or {}

    def get_daily_slaughter_data(self):
        """Get daily slaughter data matching Excel format"""
        from django.conf import settings
        from django.db import connection
        from django.db.models import Prefetch
        from django.utils import timezone

        from processing.models import Animal, WeightLog

        tenant = getattr(connection, "tenant", None)
        tz_name = getattr(tenant, "timezone", "") or getattr(settings, "TIME_ZONE", "UTC")
        try:
            tenant_tz = ZoneInfo(tz_name)
        except Exception:
            tenant_tz = timezone.get_current_timezone()

        start_dt = timezone.make_aware(datetime.combine(self.start_date, time.min), tenant_tz)
        end_dt = timezone.make_aware(datetime.combine(self.end_date + timedelta(days=1), time.min), tenant_tz)
        statuses = self.filters.get("statuses") or ["carcass_ready", "disassembled", "packaged", "delivered"]

        animals = (
            Animal.objects.filter(
                is_active=True,
                slaughter_date__gte=start_dt,
                slaughter_date__lt=end_dt,
                status__in=statuses,
                slaughter_order__is_active=True,
            )
            .select_related(
                "slaughter_order",
                "slaughter_order__client",
                "slaughter_order__destination_client",
                "cattle_details",
                "sheep_details",
                "goat_details",
                "lamb_details",
                "oglak_details",
                "calf_details",
                "heifer_details",
                "beef_details",
            )
            .prefetch_related(
                Prefetch(
                    "individual_weight_logs",
                    queryset=WeightLog.objects.filter(
                        weight_type__in=["live_weight", "hot_carcass_weight", "cold_carcass_weight"]
                    ),
                    to_attr="_prefetched_weight_logs",
                )
            )
        )

        animal_types = self.filters.get("animal_types") or []
        if animal_types:
            animals = animals.filter(animal_type__in=animal_types)

        client_ids = self.filters.get("client_ids") or []
        if client_ids:
            animals = animals.filter(slaughter_order__client_id__in=client_ids)

        destination_client_ids = self.filters.get("destination_client_ids") or []
        if destination_client_ids:
            animals = animals.filter(slaughter_order__destination_client_id__in=destination_client_ids)

        service_package_ids = self.filters.get("service_package_ids") or []
        if service_package_ids:
            animals = animals.filter(slaughter_order__service_package_id__in=service_package_ids)

        if self.filters.get("include_walkins") is False:
            animals = animals.filter(slaughter_order__client__isnull=False)

        daily_data = []

        for animal in animals:
            live_weight = self._get_weight(animal, "live_weight")
            hot_carcass_weight = self._get_weight(animal, "hot_carcass_weight")
            leather_weight = animal.leather_weight_kg or 0

            offal_status, bowels_status = self._get_offal_bowels_status(animal)
            destination = self._get_destination(animal)
            client_name = self._get_client_name(animal)

            turkish_type = self._get_turkish_animal_type(animal.animal_type)
            is_small_animal = turkish_type in ["KUZU", "OGLAK", "KECI", "KOYUN"]

            daily_data.append(
                {
                    "identification_tag": "" if is_small_animal else (animal.identification_tag or ""),
                    "client_name": client_name,
                    "quantity": 1,
                    "animal_type": turkish_type,
                    "live_weight": live_weight,
                    "hot_carcass_weight": hot_carcass_weight,
                    "offal_status": offal_status,
                    "bowels_status": bowels_status,
                    "leather_weight": leather_weight,
                    "sakatat_weight": 1.0 if offal_status == "SAĞLAM" else 0.0,
                    "destination": destination,
                    "description": "",
                }
            )

        return self._aggregate_identical_records(daily_data)

    def get_daily_summary_totals(self):
        """Get summary totals by animal type"""
        return self._compute_summary(self.get_daily_slaughter_data())

    def _compute_summary(self, daily_data):
        """Compute per-type summary totals from already-fetched daily_data."""
        summary = {
            "buyukbas": {"kesim": 0, "deri": 0, "bagirsak": 0, "sakatat": 0},
            "kuzu": {"kesim": 0, "deri": 0, "bagirsak": 0, "sakatat": 0},
            "oglak": {"kesim": 0, "deri": 0, "bagirsak": 0, "sakatat": 0},
            "koyun": {"kesim": 0, "deri": 0, "bagirsak": 0, "sakatat": 0},
            "keci": {"kesim": 0, "deri": 0, "bagirsak": 0, "sakatat": 0},
        }
        for item in daily_data:
            animal_type = item["animal_type"].upper()
            leather_weight = item["leather_weight"] or 0
            quantity = item["quantity"]
            if animal_type in ["SIGIR", "DUVE", "DANA"]:
                summary["buyukbas"]["kesim"] += quantity
                summary["buyukbas"]["deri"] += leather_weight
                if item["bowels_status"] == "SAĞLAM":
                    summary["buyukbas"]["bagirsak"] += quantity
                if item.get("sakatat_weight", 0) == 1.0:
                    summary["buyukbas"]["sakatat"] += quantity
            elif animal_type == "KUZU":
                summary["kuzu"]["kesim"] += quantity
                summary["kuzu"]["deri"] += leather_weight
                if item["bowels_status"] == "SAĞLAM":
                    summary["kuzu"]["bagirsak"] += quantity
                if item.get("sakatat_weight", 0) == 1.0:
                    summary["kuzu"]["sakatat"] += quantity
            elif animal_type == "OGLAK":
                summary["oglak"]["kesim"] += quantity
                summary["oglak"]["deri"] += leather_weight
                if item["bowels_status"] == "SAĞLAM":
                    summary["oglak"]["bagirsak"] += quantity
                if item.get("sakatat_weight", 0) == 1.0:
                    summary["oglak"]["sakatat"] += quantity
            elif animal_type == "KOYUN":
                summary["koyun"]["kesim"] += quantity
                summary["koyun"]["deri"] += leather_weight
                if item["bowels_status"] == "SAĞLAM":
                    summary["koyun"]["bagirsak"] += quantity
                if item.get("sakatat_weight", 0) == 1.0:
                    summary["koyun"]["sakatat"] += quantity
            elif animal_type == "KECI":
                summary["keci"]["kesim"] += quantity
                summary["keci"]["deri"] += leather_weight
                if item.get("sakatat_weight", 0) == 1.0:
                    summary["keci"]["sakatat"] += quantity
        return summary

    def get_status_breakdown_slaughter_period(self):
        """Count animals by workflow status where slaughter_date is in the selected range."""
        from processing.models import Animal

        rows = (
            Animal.objects.filter(slaughter_date__date__range=[self.start_date, self.end_date])
            .values("status")
            .annotate(n=Count("id"))
        )
        return {row["status"]: row["n"] for row in rows}

    def get_wip_intake_by_received_date(self):
        """Headcount in early stages for animals whose received_date falls in the range (backlog / WIP)."""
        from processing.models import Animal

        rows = (
            Animal.objects.filter(
                received_date__date__range=[self.start_date, self.end_date],
                status__in=["received", "slaughtered"],
            )
            .values("status")
            .annotate(n=Count("id"))
        )
        return {row["status"]: row["n"] for row in rows}

    @staticmethod
    def _compute_yield_stats(daily_data):
        """Hot vs live yield % (100 * hot/live) for rows with both weights > 0."""
        ratios = []
        for item in daily_data:
            live = float(item.get("live_weight") or 0)
            hot = float(item.get("hot_carcass_weight") or 0)
            if live > 0 and hot > 0:
                ratios.append(100.0 * hot / live)
        if not ratios:
            return {
                "sample_count": 0,
                "mean_yield_pct": None,
                "median_yield_pct": None,
            }
        return {
            "sample_count": len(ratios),
            "mean_yield_pct": round(statistics.mean(ratios), 2),
            "median_yield_pct": round(statistics.median(ratios), 2),
        }

    @staticmethod
    def _rollup_by_client(daily_data):
        """Aggregate headcount and hot carcass kg by client (firm) name."""
        agg = defaultdict(lambda: {"headcount": 0, "hot_carcass_kg": 0.0})
        for item in daily_data:
            key = (item.get("client_name") or "").strip() or "—"
            q = int(item.get("quantity") or 0)
            agg[key]["headcount"] += q
            agg[key]["hot_carcass_kg"] += float(item.get("hot_carcass_weight") or 0)
        return dict(sorted(agg.items(), key=lambda kv: (-kv[1]["headcount"], kv[0])))

    @staticmethod
    def _rollup_by_animal_type(daily_data):
        """Aggregate headcount and hot carcass kg by Turkish animal type label."""
        agg = defaultdict(lambda: {"headcount": 0, "hot_carcass_kg": 0.0})
        for item in daily_data:
            key = item.get("animal_type") or "—"
            q = int(item.get("quantity") or 0)
            agg[key]["headcount"] += q
            agg[key]["hot_carcass_kg"] += float(item.get("hot_carcass_weight") or 0)
        return dict(sorted(agg.items(), key=lambda kv: (-kv[1]["headcount"], kv[0])))

    def get_all_data(self):
        """Get all data for report generation"""
        daily_data = self.get_daily_slaughter_data()
        base = {
            "date": self.start_date.strftime("%Y-%m-%d"),
            "start_date": self.start_date.strftime("%Y-%m-%d"),
            "end_date": self.end_date.strftime("%Y-%m-%d"),
            "daily_data": daily_data,
            "summary": self._compute_summary(daily_data),
            "total_animals": sum(item["quantity"] for item in daily_data),  # Total count of all animals
            "total_live_weight": sum(
                item["live_weight"] for item in daily_data
            ),  # Already multiplied by quantity in aggregation
            "total_hot_carcass_weight": sum(
                item["hot_carcass_weight"] for item in daily_data
            ),  # Already multiplied by quantity in aggregation
            "total_leather_weight": sum(
                item["leather_weight"] for item in daily_data
            ),  # Already multiplied by quantity in aggregation
            "status_breakdown_slaughter_period": self.get_status_breakdown_slaughter_period(),
            "wip_intake_by_received_date": self.get_wip_intake_by_received_date(),
            "yield_stats": self._compute_yield_stats(daily_data),
            "by_client": self._rollup_by_client(daily_data),
            "by_animal_type": self._rollup_by_animal_type(daily_data),
        }
        rt = self.filters.get("report_type", "daily_slaughter")
        if rt == "cold_shrinkage":
            base["cold_shrinkage"] = self.get_cold_shrinkage_data()
        elif rt == "cut_yield_analysis":
            base["cut_yield"] = self.get_cut_yield_data()
        elif rt == "pipeline_time_analysis":
            base["pipeline_time"] = self.get_pipeline_time_data()
        elif rt == "byproduct_offal_summary":
            base["byproduct_offal"] = self.get_byproduct_offal_data()
        elif rt == "client_activity_summary":
            base["client_activity"] = self.get_client_activity_data()
        return base

    def get_cold_shrinkage_data(self):
        """Cold vs hot carcass weight — fire (shrinkage) metrics."""
        from inventory.models import Carcass

        carcasses = Carcass.objects.filter(
            animal__slaughter_date__date__range=[self.start_date, self.end_date],
            cold_carcass_weight__isnull=False,
            hot_carcass_weight__gt=0,
        ).select_related("animal", "animal__slaughter_order", "animal__slaughter_order__client")

        rows = []
        ratios = []
        by_type: dict = defaultdict(list)

        for c in carcasses:
            hot = float(c.hot_carcass_weight)
            cold = float(c.cold_carcass_weight)
            shrink_pct = round(100.0 * (hot - cold) / hot, 2) if hot > 0 else None
            turkish_type = self._get_turkish_animal_type(c.animal.animal_type)
            rows.append({
                "identification_tag": c.animal.identification_tag or "",
                "animal_type": turkish_type,
                "client_name": self._get_client_name(c.animal),
                "hot_carcass_weight": hot,
                "cold_carcass_weight": cold,
                "shrinkage_kg": round(hot - cold, 2),
                "shrinkage_pct": shrink_pct,
            })
            if shrink_pct is not None:
                ratios.append(shrink_pct)
                by_type[turkish_type].append(shrink_pct)

        type_summary = {
            t: {
                "sample_count": len(vals),
                "mean_pct": round(statistics.mean(vals), 2),
                "median_pct": round(statistics.median(vals), 2),
            }
            for t, vals in by_type.items()
        }

        return {
            "rows": rows,
            "sample_count": len(ratios),
            "mean_shrinkage_pct": round(statistics.mean(ratios), 2) if ratios else None,
            "median_shrinkage_pct": round(statistics.median(ratios), 2) if ratios else None,
            "by_animal_type": type_summary,
        }

    def get_cut_yield_data(self):
        """Disassembly cut breakdown — kg per cut type as % of hot carcass."""
        from django.db.models import Q, Sum

        from processing.models import DisassemblyCut, WeightLog

        cuts = (
            DisassemblyCut.objects.filter(
                animal__slaughter_date__date__range=[self.start_date, self.end_date]
            )
            .values("cut_name")
            .annotate(
                total_kg=Sum("weight_kg"),
                cut_count=Count("id"),
                scale_synced=Count("source_event", filter=Q(source_event__isnull=False)),
            )
            .order_by("-total_kg")
        )

        total_hot = (
            WeightLog.objects.filter(
                weight_type="hot_carcass_weight",
                is_group_weight=False,
                animal__slaughter_date__date__range=[self.start_date, self.end_date],
            ).aggregate(total=Sum("weight"))["total"]
            or 0
        )
        total_hot_float = float(total_hot)

        cut_rows = []
        for cut in cuts:
            kg = float(cut["total_kg"] or 0)
            cut_rows.append({
                "cut_name": cut["cut_name"],
                "total_kg": round(kg, 2),
                "cut_count": cut["cut_count"],
                "scale_synced": cut["scale_synced"],
                "pct_of_hot": round(100.0 * kg / total_hot_float, 2) if total_hot_float > 0 else None,
            })

        return {
            "cuts": cut_rows,
            "total_hot_carcass_kg": round(total_hot_float, 2),
        }

    def get_pipeline_time_data(self):
        """Processing stage duration — intake→slaughter and disassembly session times."""
        from django.db.models import DurationField, ExpressionWrapper, F

        from processing.models import Animal as AnimalModel
        from scales.models import DisassemblySession

        animals = AnimalModel.objects.filter(
            slaughter_date__date__range=[self.start_date, self.end_date],
            received_date__isnull=False,
            slaughter_date__isnull=False,
        ).annotate(
            intake_lag=ExpressionWrapper(
                F("slaughter_date") - F("received_date"), output_field=DurationField()
            )
        )

        intake_hours = [
            a.intake_lag.total_seconds() / 3600
            for a in animals
            if a.intake_lag and a.intake_lag.total_seconds() >= 0
        ]

        sessions = DisassemblySession.objects.filter(
            started_at__date__range=[self.start_date, self.end_date],
            ended_at__isnull=False,
            status__in=["completed", "auto_closed"],
        )
        session_minutes = [
            (s.ended_at - s.started_at).total_seconds() / 60
            for s in sessions
            if s.ended_at > s.started_at
        ]

        def _stats(vals):
            if not vals:
                return {"sample_count": 0, "mean": None, "median": None}
            return {
                "sample_count": len(vals),
                "mean": round(statistics.mean(vals), 2),
                "median": round(statistics.median(vals), 2),
            }

        return {
            "intake_to_slaughter_hours": _stats(intake_hours),
            "disassembly_session_minutes": _stats(session_minutes),
            "animals_by_status": self.get_status_breakdown_slaughter_period(),
        }

    def get_byproduct_offal_data(self):
        """Offal and by-product weight breakdown by type and disposition."""
        from django.db.models import Sum

        from inventory.models import ByProduct, Offal

        offal_rows = list(
            Offal.objects.filter(
                animal__slaughter_date__date__range=[self.start_date, self.end_date]
            )
            .values("offal_type", "disposition")
            .annotate(total_weight=Sum("weight"), count=Count("id"))
            .order_by("offal_type", "disposition")
        )

        byproduct_rows = list(
            ByProduct.objects.filter(
                animal__slaughter_date__date__range=[self.start_date, self.end_date]
            )
            .values("byproduct_type", "disposition")
            .annotate(total_weight=Sum("weight"), count=Count("id"))
            .order_by("byproduct_type", "disposition")
        )

        return {
            "offal_by_type": [
                {
                    "offal_type": r["offal_type"],
                    "disposition": r["disposition"],
                    "total_weight": float(r["total_weight"] or 0),
                    "count": r["count"],
                }
                for r in offal_rows
            ],
            "byproducts_by_type": [
                {
                    "byproduct_type": r["byproduct_type"],
                    "disposition": r["disposition"],
                    "total_weight": float(r["total_weight"] or 0),
                    "count": r["count"],
                }
                for r in byproduct_rows
            ],
        }

    def get_client_activity_data(self):
        """Per-client summary: order count, animal count, total hot carcass kg."""
        from django.db.models import Sum

        from reception.models import SlaughterOrder

        rows = list(
            SlaughterOrder.objects.filter(
                order_datetime__date__range=[self.start_date, self.end_date]
            )
            .values(
                "client__company_name",
                "client__contact_person",
                "client_name",
                "status",
            )
            .annotate(
                order_count=Count("id"),
                animal_count=Count("animals"),
                total_hot_carcass_kg=Sum("animals__carcass__hot_carcass_weight"),
            )
            .order_by("-animal_count")
        )

        result = []
        for r in rows:
            client_label = (
                r["client__company_name"]
                or r["client__contact_person"]
                or r["client_name"]
                or "Walk-in"
            )
            result.append({
                "client_name": client_label,
                "status": r["status"],
                "order_count": r["order_count"],
                "animal_count": r["animal_count"],
                "total_hot_carcass_kg": round(float(r["total_hot_carcass_kg"] or 0), 2),
            })
        return result

    def get_client_order_data(self, client_profile_id):
        """Per-order detail scoped to a single client — used by customer portal."""
        from processing.models import Animal as AnimalModel

        animals = (
            AnimalModel.objects.filter(
                slaughter_order__client_id=client_profile_id,
                slaughter_order__order_datetime__date__range=[self.start_date, self.end_date],
            )
            .select_related(
                "slaughter_order",
                "slaughter_order__client",
                "slaughter_order__service_package",
            )
            .prefetch_related(
                "individual_weight_logs",
                "disassembly_cuts",
                "offals",
                "by_products",
            )
        )

        orders: dict = {}
        for animal in animals:
            order = animal.slaughter_order
            if order.pk not in orders:
                orders[order.pk] = {
                    "order_no": order.slaughter_order_no,
                    "order_date": order.order_datetime.strftime("%Y-%m-%d"),
                    "status": order.status,
                    "service_package": order.service_package.name if order.service_package else "",
                    "animals": [],
                }
            turkish_type = self._get_turkish_animal_type(animal.animal_type)
            is_small = turkish_type in ["KUZU", "OGLAK", "KECI", "KOYUN"]
            offal_s, bowels_s = self._get_offal_bowels_status(animal)
            cuts = [
                {"cut_name": c.cut_name, "weight_kg": float(c.weight_kg)}
                for c in animal.disassembly_cuts.all()
            ]
            orders[order.pk]["animals"].append({
                "identification_tag": "" if is_small else (animal.identification_tag or ""),
                "animal_type": turkish_type,
                "status": animal.status,
                "live_weight": self._get_weight(animal, "live_weight"),
                "hot_carcass_weight": self._get_weight(animal, "hot_carcass_weight"),
                "leather_weight": float(animal.leather_weight_kg or 0),
                "offal_status": offal_s,
                "bowels_status": bowels_s,
                "cuts": cuts,
            })
        return list(orders.values())

    def _get_weight(self, animal, weight_type):
        """Get specific weight for animal, using prefetched logs when available."""
        try:
            prefetched = getattr(animal, "_prefetched_weight_logs", None)
            if prefetched is not None:
                weight_log = next((log for log in prefetched if log.weight_type == weight_type), None)
            else:
                weight_log = animal.individual_weight_logs.filter(weight_type=weight_type).first()
            return float(weight_log.weight) if weight_log else 0
        except (TypeError, ValueError, AttributeError):
            return 0

    def _get_offal_bowels_status(self, animal):
        """Get offal and bowels status from detail models"""
        offal_status = "SAĞLAM"  # Default
        bowels_status = "SAĞLAM"  # Default

        # Check detail models for status based on animal type
        details = None
        if animal.animal_type == "cattle" and hasattr(animal, "cattle_details"):
            details = animal.cattle_details
        elif animal.animal_type == "sheep" and hasattr(animal, "sheep_details"):
            details = animal.sheep_details
        elif animal.animal_type == "goat" and hasattr(animal, "goat_details"):
            details = animal.goat_details
        elif animal.animal_type == "lamb" and hasattr(animal, "lamb_details"):
            details = animal.lamb_details
        elif animal.animal_type == "oglak" and hasattr(animal, "oglak_details"):
            details = animal.oglak_details
        elif animal.animal_type == "calf" and hasattr(animal, "calf_details"):
            details = animal.calf_details
        elif animal.animal_type == "heifer" and hasattr(animal, "heifer_details"):
            details = animal.heifer_details
        elif animal.animal_type == "beef" and hasattr(animal, "beef_details"):
            details = animal.beef_details

        if details:
            logger.debug(
                "Detail %s for %s — sakatat=%s bowels=%s",
                details.__class__.__name__,
                animal.identification_tag,
                getattr(details, "sakatat_status", None),
                getattr(details, "bowels_status", None),
            )

            # Set offal status based on sakatat_status
            if details.sakatat_status == 0:
                offal_status = "ATIK"
            elif details.sakatat_status == 0.5:
                offal_status = "YARIM"
            else:  # 1.0
                offal_status = "SAĞLAM"

            # Set bowels status based on bowels_status
            if details.bowels_status == 0:
                bowels_status = "BOZUK"
            elif details.bowels_status == 0.5:
                bowels_status = "YARIM"
            else:  # 1.0
                bowels_status = "SAĞLAM"
        else:
            logger.debug(
                "No detail model for %s (type=%s); defaulting offal/bowels",
                animal.identification_tag,
                animal.animal_type,
            )
            offal_status = "SAĞLAM"  # Default to good
            bowels_status = "SAĞLAM"  # Default to good

        return offal_status, bowels_status

    def _get_destination(self, animal):
        """Get destination customer"""
        # This would need to be implemented based on your business logic
        # Could be from inventory disposition or order destination
        return animal.slaughter_order.destination or ""

    def _get_client_name(self, animal):
        """Get client name from slaughter order"""
        if animal.slaughter_order.client:
            return animal.slaughter_order.client.company_name or animal.slaughter_order.client.contact_person
        else:
            return animal.slaughter_order.client_name or "Walk-in Customer"

    def _get_turkish_animal_type(self, animal_type):
        """Convert English animal type to Turkish - matches labeling system mapping"""
        type_mapping = {
            "cattle": "SIGIR",
            "sheep": "KOYUN",
            "goat": "KECI",
            "lamb": "KUZU",
            "oglak": "OGLAK",
            "calf": "BUZA",
            "heifer": "DUVE",
            "beef": "DANA",
        }
        return type_mapping.get(animal_type, animal_type.upper())

    def _aggregate_identical_records(self, daily_data):
        """
        Group identical records and sum their quantities.
        Records are considered identical only if ALL fields match:
        - client_name, animal_type, offal_status, bowels_status,
        - leather_weight, destination, description

        If ANY field differs, a separate row is created.
        Examples:
        - Different animal type (KUZU vs DANA) → separate rows
        - Different weight (25 vs 30) → separate rows
        - Different offal status (SAĞLAM vs ATIK) → separate rows
        - Different client → separate rows
        """
        # Create a dictionary to group records by their key fields
        grouped_records = {}

        for record in daily_data:
            # Create a key based on ALL fields except quantity and weights
            # For small animals (KUZU/OGLAK/KECI), omit identification_tag so same rows aggregate
            # For others, include identification_tag to keep rows unique per animal
            if record["animal_type"] in ["KUZU", "OGLAK", "KECI", "KOYUN"]:
                key = (
                    record["client_name"],
                    record["animal_type"],
                    record["live_weight"],
                    record["hot_carcass_weight"],
                    record["offal_status"],
                    record["bowels_status"],
                    record["leather_weight"],
                    record["sakatat_weight"],
                    record["destination"],
                    record["description"],
                )
            else:
                key = (
                    record.get("identification_tag", ""),
                    record["client_name"],
                    record["animal_type"],
                    record["live_weight"],
                    record["hot_carcass_weight"],
                    record["offal_status"],
                    record["bowels_status"],
                    record["leather_weight"],
                    record["sakatat_weight"],
                    record["destination"],
                    record["description"],
                )

            if key in grouped_records:
                # If record exists, add to quantity
                existing = grouped_records[key]
                existing["quantity"] += record["quantity"]
                # Weights should be the same for identical records, so no averaging needed
            else:
                # Create new record - this happens when ANY field is different
                grouped_records[key] = record.copy()

        # After grouping, multiply weights by total quantity for each record
        for record in grouped_records.values():
            quantity = record["quantity"]
            if quantity > 1:
                # Multiply weights by quantity to show total weights
                record["live_weight"] = record["live_weight"] * quantity
                record["hot_carcass_weight"] = record["hot_carcass_weight"] * quantity
                record["leather_weight"] = record["leather_weight"] * quantity

        # Convert back to list
        return list(grouped_records.values())


class OperationsInsightService:
    """
    Real-time operational and quality insight service for the reporting dashboard.

    Powers two endpoints behind /reporting/api/:
      - ops_kpis:        live WIP, throughput, bottleneck dwell, yield trend, device health
      - quality_insight: per-type baselines, per-animal outliers, per-client scorecard
    """

    TERMINAL_STATUSES = ("delivered", "returned", "disposed")
    OPEN_STATUSES = ("received", "slaughtered", "carcass_ready", "disassembled", "packaged")

    _ANIMAL_TYPE_TURKISH = {
        "cattle": "SIGIR",
        "sheep": "KOYUN",
        "goat": "KECI",
        "lamb": "KUZU",
        "oglak": "OGLAK",
        "calf": "BUZA",
        "heifer": "DUVE",
        "beef": "DANA",
    }

    _DETAIL_REL_BY_TYPE = {
        "cattle": "cattle_details",
        "sheep": "sheep_details",
        "goat": "goat_details",
        "lamb": "lamb_details",
        "oglak": "oglak_details",
        "calf": "calf_details",
        "heifer": "heifer_details",
        "beef": "beef_details",
    }

    # ---------- helpers ----------

    @staticmethod
    def _tenant_tz():
        from django.conf import settings
        from django.db import connection
        from django.utils import timezone as djtz

        tenant = getattr(connection, "tenant", None)
        tz_name = getattr(tenant, "timezone", "") or getattr(settings, "TIME_ZONE", "UTC")
        try:
            return ZoneInfo(tz_name)
        except Exception:
            return djtz.get_current_timezone()

    @staticmethod
    def _now():
        from django.utils import timezone as djtz

        return djtz.now()

    @staticmethod
    def _percentile(values, pct):
        if not values:
            return None
        s = sorted(values)
        k = (len(s) - 1) * (pct / 100.0)
        f = int(k)
        c = min(f + 1, len(s) - 1)
        if f == c:
            return round(float(s[f]), 2)
        return round(float(s[f] + (s[c] - s[f]) * (k - f)), 2)

    @staticmethod
    def _stats(values):
        if not values:
            return {"n": 0, "mean": None, "p50": None, "p90": None}
        return {
            "n": len(values),
            "mean": round(statistics.mean(values), 2),
            "p50": OperationsInsightService._percentile(values, 50),
            "p90": OperationsInsightService._percentile(values, 90),
        }

    # ---------- Idea 1: Live Ops KPIs ----------

    def get_live_ops_kpis(self):
        """One call for the live-ops panel — avoids N endpoints."""
        return {
            "generated_at": self._now().isoformat(),
            "timezone": str(self._tenant_tz()),
            "today": self._today_throughput(),
            "wip_snapshot": self._wip_snapshot(),
            "bottleneck": self._bottleneck_stats(),
            "oldest_wip": self._oldest_wip_by_status(),
            "yield_trend": self._yield_trend(days=14),
            "scale_health": self._scale_health(),
        }

    def _today_throughput(self):
        from processing.models import Animal

        tz = self._tenant_tz()
        now = self._now().astimezone(tz)
        today = now.date()
        slaughtered_today = (
            Animal.objects.filter(
                slaughter_date__date=today,
                status__in=["slaughtered", "carcass_ready", "disassembled", "packaged", "delivered"],
            ).count()
        )
        received_today = Animal.objects.filter(received_date__date=today).count()
        delivered_today = Animal.objects.filter(status="delivered", updated_at__date=today).count()
        by_status = {
            row["status"]: row["n"]
            for row in Animal.objects.filter(slaughter_date__date=today)
            .values("status")
            .annotate(n=Count("id"))
        }
        return {
            "date": today.isoformat(),
            "received": received_today,
            "slaughtered": slaughtered_today,
            "delivered": delivered_today,
            "by_status_slaughtered_today": by_status,
        }

    def _wip_snapshot(self):
        """Headcount currently in each open status (date-independent)."""
        from processing.models import Animal

        rows = (
            Animal.objects.filter(is_active=True, status__in=self.OPEN_STATUSES)
            .values("status")
            .annotate(n=Count("id"))
        )
        snapshot = {s: 0 for s in self.OPEN_STATUSES}
        for row in rows:
            snapshot[row["status"]] = row["n"]
        snapshot["total_open"] = sum(snapshot.values())
        return snapshot

    def _bottleneck_stats(self):
        """Rolling 24h dwell stats: intake→slaughter, disassembly session duration, carcass_ready dwell."""
        from django.utils import timezone as djtz

        from processing.models import Animal
        from scales.models import DisassemblySession

        now = djtz.now()
        window_start = now - timedelta(hours=24)

        intake_animals = Animal.objects.filter(
            slaughter_date__gte=window_start,
            received_date__isnull=False,
            slaughter_date__isnull=False,
        ).only("id", "received_date", "slaughter_date")
        intake_hours = [
            (a.slaughter_date - a.received_date).total_seconds() / 3600
            for a in intake_animals
            if a.slaughter_date and a.received_date and a.slaughter_date >= a.received_date
        ]

        sessions = DisassemblySession.objects.filter(
            ended_at__isnull=False,
            ended_at__gte=window_start,
            status__in=["completed", "auto_closed"],
        ).only("id", "started_at", "ended_at")
        session_minutes = [
            (s.ended_at - s.started_at).total_seconds() / 60
            for s in sessions
            if s.ended_at and s.started_at and s.ended_at > s.started_at
        ]

        carcass_dwell = Animal.objects.filter(
            status="carcass_ready", is_active=True, updated_at__isnull=False
        ).only("id", "updated_at")
        carcass_hours = [
            (now - a.updated_at).total_seconds() / 3600
            for a in carcass_dwell
            if a.updated_at
        ]

        stats = {
            "intake_to_slaughter_hours": self._stats(intake_hours),
            "disassembly_session_minutes": self._stats(session_minutes),
            "carcass_ready_dwell_hours": self._stats(carcass_hours),
        }

        candidates = [
            ("intake_to_slaughter_hours", "Intake → Slaughter"),
            ("disassembly_session_minutes", "Disassembly session"),
            ("carcass_ready_dwell_hours", "Carcass-ready dwell"),
        ]
        worst_key = None
        worst_p90 = -1.0
        for key, _label in candidates:
            p90 = stats[key]["p90"]
            if p90 is not None and p90 > worst_p90:
                worst_p90 = p90
                worst_key = key
        stats["worst_stage"] = worst_key
        stats["worst_stage_label"] = dict(candidates).get(worst_key) if worst_key else None
        return stats

    def _oldest_wip_by_status(self):
        """For each open status, the longest-waiting animal (by updated_at)."""
        from processing.models import Animal

        now = self._now()
        out = {}
        for status in self.OPEN_STATUSES:
            oldest = (
                Animal.objects.filter(is_active=True, status=status)
                .order_by("updated_at")
                .only("id", "identification_tag", "animal_type", "updated_at")
                .first()
            )
            if oldest and oldest.updated_at:
                out[status] = {
                    "identification_tag": oldest.identification_tag or "",
                    "animal_type": self._ANIMAL_TYPE_TURKISH.get(oldest.animal_type, oldest.animal_type),
                    "hours_in_status": round((now - oldest.updated_at).total_seconds() / 3600, 1),
                }
            else:
                out[status] = None
        return out

    def _yield_trend(self, days=14):
        """Per-day mean dressing % (hot/live * 100) from WeightLog, last N days."""
        from django.db.models import F
        from django.utils import timezone as djtz

        from processing.models import Animal, WeightLog

        now = djtz.now()
        start = (now - timedelta(days=days - 1)).date()
        animals = (
            Animal.objects.filter(
                slaughter_date__date__gte=start,
                slaughter_date__isnull=False,
            )
            .only("id", "slaughter_date")
        )
        animal_map = {a.id: a.slaughter_date.date() for a in animals}
        if not animal_map:
            return []
        logs = WeightLog.objects.filter(
            animal_id__in=list(animal_map.keys()),
            weight_type__in=["live_weight", "hot_carcass_weight"],
            is_group_weight=False,
        ).values("animal_id", "weight_type", "weight")

        per_animal = defaultdict(dict)
        for row in logs:
            per_animal[row["animal_id"]][row["weight_type"]] = float(row["weight"] or 0)

        per_day = defaultdict(list)
        for aid, weights in per_animal.items():
            live = weights.get("live_weight") or 0
            hot = weights.get("hot_carcass_weight") or 0
            if live > 0 and hot > 0:
                day = animal_map[aid]
                per_day[day].append(100.0 * hot / live)

        out = []
        d = start
        today = now.date()
        while d <= today:
            vals = per_day.get(d, [])
            out.append({
                "date": d.isoformat(),
                "n": len(vals),
                "mean_pct": round(statistics.mean(vals), 2) if vals else None,
            })
            d += timedelta(days=1)
        return out

    def _scale_health(self):
        """Device + edge health summary, plus pending orphan batches."""
        try:
            from scales.models import EdgeDevice, OrphanedBatch, ScaleDevice
        except Exception:
            return {"available": False}

        try:
            status_counts = {
                row["status"]: row["n"]
                for row in ScaleDevice.objects.values("status").annotate(n=Count("id"))
            }
        except Exception:
            status_counts = {}
        try:
            total = sum(status_counts.values())
            offline_ish = sum(
                status_counts.get(k, 0) for k in ("stale", "disconnected", "unknown")
            )
            pending_batches = OrphanedBatch.objects.filter(status="pending").count()
            edges_online = EdgeDevice.objects.filter(is_online=True).count()
            edges_total = EdgeDevice.objects.count()
        except Exception:
            total = 0
            offline_ish = 0
            pending_batches = 0
            edges_online = 0
            edges_total = 0

        return {
            "available": True,
            "scale_total": total,
            "scale_status": status_counts,
            "scale_offline_like": offline_ish,
            "pending_orphan_batches": pending_batches,
            "edges_online": edges_online,
            "edges_total": edges_total,
        }

    # ---------- Idea 2: Quality & Loss insight ----------

    def get_quality_insight(self, *, window_days=90, start=None, end=None, sigma=2.0):
        """
        Returns yield/shrinkage baselines, flagged outliers in [start,end], and a per-client scorecard.
        """
        from django.utils import timezone as djtz

        now = djtz.now()
        today = now.date()
        if end is None:
            end = today
        if start is None:
            start = end - timedelta(days=6)  # default = last 7 days

        baselines_yield, samples_yield = self._yield_baselines(window_days=window_days)
        baselines_shrink, samples_shrink = self._shrinkage_baselines(window_days=window_days)

        outliers = self._yield_outliers(
            start=start,
            end=end,
            baselines_yield=baselines_yield,
            baselines_shrink=baselines_shrink,
            sigma=sigma,
        )
        scorecard = self._client_scorecard(start=start, end=end)

        return {
            "generated_at": now.isoformat(),
            "window_days": window_days,
            "range": {"start": start.isoformat(), "end": end.isoformat()},
            "sigma_threshold": sigma,
            "baselines": {
                "yield_pct_by_type": baselines_yield,
                "shrinkage_pct_by_type": baselines_shrink,
                "yield_sample_count": samples_yield,
                "shrinkage_sample_count": samples_shrink,
            },
            "outliers": outliers,
            "client_scorecard": scorecard,
        }

    def _yield_baselines(self, window_days=90):
        """Mean/stdev yield % per animal_type across a rolling window."""
        from django.utils import timezone as djtz

        from processing.models import Animal, WeightLog

        start = djtz.now() - timedelta(days=window_days)
        animals = Animal.objects.filter(
            slaughter_date__gte=start, slaughter_date__isnull=False
        ).only("id", "animal_type")
        type_by_id = {a.id: a.animal_type for a in animals}
        if not type_by_id:
            return {}, 0

        logs = WeightLog.objects.filter(
            animal_id__in=list(type_by_id.keys()),
            weight_type__in=["live_weight", "hot_carcass_weight"],
            is_group_weight=False,
        ).values("animal_id", "weight_type", "weight")
        per_animal = defaultdict(dict)
        for row in logs:
            per_animal[row["animal_id"]][row["weight_type"]] = float(row["weight"] or 0)

        by_type = defaultdict(list)
        for aid, weights in per_animal.items():
            live = weights.get("live_weight") or 0
            hot = weights.get("hot_carcass_weight") or 0
            if live > 0 and hot > 0:
                by_type[type_by_id[aid]].append(100.0 * hot / live)

        baselines = {}
        total = 0
        for t, vals in by_type.items():
            if len(vals) < 3:
                continue
            mean = statistics.mean(vals)
            stdev = statistics.pstdev(vals) if len(vals) >= 2 else 0.0
            baselines[self._ANIMAL_TYPE_TURKISH.get(t, t)] = {
                "raw_type": t,
                "n": len(vals),
                "mean_pct": round(mean, 2),
                "stdev_pct": round(stdev, 2),
            }
            total += len(vals)
        return baselines, total

    def _shrinkage_baselines(self, window_days=90):
        """Mean/stdev cold-shrinkage % per animal_type across a rolling window."""
        from django.utils import timezone as djtz

        from inventory.models import Carcass

        start = djtz.now() - timedelta(days=window_days)
        carcasses = Carcass.objects.filter(
            animal__slaughter_date__gte=start,
            cold_carcass_weight__isnull=False,
            hot_carcass_weight__gt=0,
        ).select_related("animal").only(
            "hot_carcass_weight",
            "cold_carcass_weight",
            "animal__animal_type",
        )

        by_type = defaultdict(list)
        for c in carcasses:
            hot = float(c.hot_carcass_weight or 0)
            cold = float(c.cold_carcass_weight or 0)
            if hot > 0 and cold >= 0:
                by_type[c.animal.animal_type].append(100.0 * (hot - cold) / hot)

        baselines = {}
        total = 0
        for t, vals in by_type.items():
            if len(vals) < 3:
                continue
            mean = statistics.mean(vals)
            stdev = statistics.pstdev(vals) if len(vals) >= 2 else 0.0
            baselines[self._ANIMAL_TYPE_TURKISH.get(t, t)] = {
                "raw_type": t,
                "n": len(vals),
                "mean_pct": round(mean, 2),
                "stdev_pct": round(stdev, 2),
            }
            total += len(vals)
        return baselines, total

    def _yield_outliers(self, *, start, end, baselines_yield, baselines_shrink, sigma):
        """
        Return per-animal flags for animals slaughtered in [start,end] whose dressing % or
        shrinkage % falls outside baseline mean ± sigma*stdev.
        """
        from processing.models import Animal, WeightLog

        animals = (
            Animal.objects.filter(
                slaughter_date__date__range=[start, end],
                is_active=True,
            )
            .select_related("slaughter_order", "slaughter_order__client")
            .only(
                "id",
                "identification_tag",
                "animal_type",
                "slaughter_date",
                "slaughter_order__client__company_name",
                "slaughter_order__client__contact_person",
                "slaughter_order__client_name",
            )
        )
        animal_ids = [a.id for a in animals]
        if not animal_ids:
            return []

        logs = WeightLog.objects.filter(
            animal_id__in=animal_ids,
            weight_type__in=["live_weight", "hot_carcass_weight"],
            is_group_weight=False,
        ).values("animal_id", "weight_type", "weight")
        per_animal = defaultdict(dict)
        for row in logs:
            per_animal[row["animal_id"]][row["weight_type"]] = float(row["weight"] or 0)

        try:
            from inventory.models import Carcass

            carcasses = Carcass.objects.filter(
                animal_id__in=animal_ids, cold_carcass_weight__isnull=False
            ).only("animal_id", "hot_carcass_weight", "cold_carcass_weight")
            carcass_map = {
                c.animal_id: (float(c.hot_carcass_weight or 0), float(c.cold_carcass_weight or 0))
                for c in carcasses
            }
        except Exception:
            carcass_map = {}

        outliers = []
        for a in animals:
            client_label = (
                (a.slaughter_order.client.company_name if a.slaughter_order.client else None)
                or (a.slaughter_order.client.contact_person if a.slaughter_order.client else None)
                or a.slaughter_order.client_name
                or "Walk-in"
            )
            weights = per_animal.get(a.id, {})
            live = weights.get("live_weight") or 0
            hot = weights.get("hot_carcass_weight") or 0

            turkish = self._ANIMAL_TYPE_TURKISH.get(a.animal_type, a.animal_type)
            reasons = []

            if live > 0 and hot > 0:
                yield_pct = 100.0 * hot / live
                base = baselines_yield.get(turkish)
                if base and base["stdev_pct"] > 0:
                    z = (yield_pct - base["mean_pct"]) / base["stdev_pct"]
                    if abs(z) >= sigma:
                        reasons.append({
                            "metric": "yield_pct",
                            "value": round(yield_pct, 2),
                            "baseline_mean": base["mean_pct"],
                            "baseline_stdev": base["stdev_pct"],
                            "z": round(z, 2),
                        })

            if a.id in carcass_map:
                hot_c, cold_c = carcass_map[a.id]
                if hot_c > 0:
                    shrink_pct = 100.0 * (hot_c - cold_c) / hot_c
                    base = baselines_shrink.get(turkish)
                    if base and base["stdev_pct"] > 0:
                        z = (shrink_pct - base["mean_pct"]) / base["stdev_pct"]
                        if abs(z) >= sigma:
                            reasons.append({
                                "metric": "shrinkage_pct",
                                "value": round(shrink_pct, 2),
                                "baseline_mean": base["mean_pct"],
                                "baseline_stdev": base["stdev_pct"],
                                "z": round(z, 2),
                            })

            if reasons:
                outliers.append({
                    "animal_id": str(a.id),
                    "identification_tag": a.identification_tag or "",
                    "animal_type": turkish,
                    "client_name": client_label,
                    "slaughter_date": a.slaughter_date.date().isoformat() if a.slaughter_date else None,
                    "reasons": reasons,
                })

        outliers.sort(
            key=lambda o: max((abs(r["z"]) for r in o["reasons"]), default=0), reverse=True
        )
        return outliers[:50]  # cap payload

    def _client_scorecard(self, *, start, end):
        """Per-client: mean yield %, mean shrinkage %, condemnation rate, headcount."""
        from django.db.models import Avg, Count as DjCount, Q

        from inventory.models import ByProduct, Carcass, Offal
        from processing.models import Animal, WeightLog

        animals = (
            Animal.objects.filter(
                slaughter_date__date__range=[start, end],
                is_active=True,
            )
            .select_related("slaughter_order", "slaughter_order__client")
            .only(
                "id",
                "animal_type",
                "slaughter_order__client__company_name",
                "slaughter_order__client__contact_person",
                "slaughter_order__client_name",
            )
        )
        if not animals:
            return []

        aid_to_client = {}
        for a in animals:
            label = (
                (a.slaughter_order.client.company_name if a.slaughter_order.client else None)
                or (a.slaughter_order.client.contact_person if a.slaughter_order.client else None)
                or a.slaughter_order.client_name
                or "Walk-in"
            )
            aid_to_client[a.id] = label

        logs = WeightLog.objects.filter(
            animal_id__in=list(aid_to_client.keys()),
            weight_type__in=["live_weight", "hot_carcass_weight"],
            is_group_weight=False,
        ).values("animal_id", "weight_type", "weight")
        per_animal = defaultdict(dict)
        for row in logs:
            per_animal[row["animal_id"]][row["weight_type"]] = float(row["weight"] or 0)

        carcass_map = {
            c.animal_id: (float(c.hot_carcass_weight or 0), float(c.cold_carcass_weight or 0))
            for c in Carcass.objects.filter(
                animal_id__in=list(aid_to_client.keys()),
                cold_carcass_weight__isnull=False,
            ).only("animal_id", "hot_carcass_weight", "cold_carcass_weight")
        }

        disposed_offal_ids = set(
            Offal.objects.filter(
                animal_id__in=list(aid_to_client.keys()), disposition="disposed"
            )
            .values_list("animal_id", flat=True)
            .distinct()
        )
        disposed_bp_ids = set(
            ByProduct.objects.filter(
                animal_id__in=list(aid_to_client.keys()), disposition="disposed"
            )
            .values_list("animal_id", flat=True)
            .distinct()
        )
        condemn_from_detail = set()
        for rel in self._DETAIL_REL_BY_TYPE.values():
            try:
                qs = (
                    Animal.objects.filter(
                        id__in=list(aid_to_client.keys()),
                    )
                    .filter(**{f"{rel}__sakatat_status__lt": 1.0})
                    .values_list("id", flat=True)
                )
                condemn_from_detail.update(qs)
                qs = (
                    Animal.objects.filter(
                        id__in=list(aid_to_client.keys()),
                    )
                    .filter(**{f"{rel}__bowels_status__lt": 1.0})
                    .values_list("id", flat=True)
                )
                condemn_from_detail.update(qs)
            except Exception:
                pass

        condemn_ids = disposed_offal_ids | disposed_bp_ids | condemn_from_detail

        client_bucket = defaultdict(lambda: {
            "headcount": 0,
            "yields": [],
            "shrinks": [],
            "condemned": 0,
        })
        for aid, client in aid_to_client.items():
            b = client_bucket[client]
            b["headcount"] += 1
            weights = per_animal.get(aid, {})
            live = weights.get("live_weight") or 0
            hot = weights.get("hot_carcass_weight") or 0
            if live > 0 and hot > 0:
                b["yields"].append(100.0 * hot / live)
            if aid in carcass_map:
                hot_c, cold_c = carcass_map[aid]
                if hot_c > 0:
                    b["shrinks"].append(100.0 * (hot_c - cold_c) / hot_c)
            if aid in condemn_ids:
                b["condemned"] += 1

        rows = []
        for client, b in client_bucket.items():
            rows.append({
                "client_name": client,
                "headcount": b["headcount"],
                "mean_yield_pct": round(statistics.mean(b["yields"]), 2) if b["yields"] else None,
                "mean_shrinkage_pct": round(statistics.mean(b["shrinks"]), 2) if b["shrinks"] else None,
                "condemnation_rate_pct": (
                    round(100.0 * b["condemned"] / b["headcount"], 2) if b["headcount"] else None
                ),
                "condemned_count": b["condemned"],
            })
        rows.sort(key=lambda r: (-(r["condemnation_rate_pct"] or 0), -r["headcount"]))
        return rows


class ExcelReportGenerator:
    """Service for generating Excel reports matching the exact format"""

    def __init__(self, report_data, template_config=None):
        self.report_data = report_data
        self.template_config = template_config or {}

    def generate_daily_slaughter_excel(self):
        """Generate Excel matching your example format"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Slaughter Report"

        # Title
        date_str = self.report_data.get("date", self.report_data.get("start_date", ""))
        ws["A1"] = f"GÜNLÜK KESİM RAPORU - {date_str}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:J1")  # Updated to J1 for 10 columns

        # Main table headers
        headers = [
            "FİRMA ÜNVANI",
            "ADET",
            "CİNSİ",
            "SICAK KARKAS",
            "HAYVAN KİMLİK NO",
            "SAKATAT",
            "BAĞIRSAK",
            "DERİ",
            "ALINAN MÜŞTERİ",
            "AÇIKLAMA",
        ]

        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Write data
        row = 4
        for item in self.report_data["daily_data"]:
            ws.cell(row=row, column=1, value=item["destination"])  # ALINAN MÜŞTERİ (first column)
            ws.cell(row=row, column=2, value=float(item["quantity"]))
            ws.cell(row=row, column=3, value=item["animal_type"])
            ws.cell(row=row, column=4, value=float(item["hot_carcass_weight"]))  # SICAK KARKAS
            ws.cell(row=row, column=5, value=item.get("identification_tag", ""))
            ws.cell(row=row, column=6, value=item["offal_status"])
            ws.cell(row=row, column=7, value=item["bowels_status"])
            ws.cell(row=row, column=8, value=float(item["leather_weight"]))
            ws.cell(row=row, column=9, value=item["client_name"])  # FİRMA ÜNVANI (before last column)
            ws.cell(row=row, column=10, value=item["description"])

            # Apply borders to data rows
            for col in range(1, 11):  # Updated to 11 for 10 columns
                ws.cell(row=row, column=col).border = thin_border

            row += 1

        # Summary section
        summary_start_row = row + 2
        ws.cell(row=summary_start_row, column=1, value="ÖZET").font = Font(bold=True, size=12)

        # Summary headers (remove goat-specific note; keep columns consistent)
        summary_headers = ["", "KESİM", "DERİ", "BAĞIRSAK"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=summary_start_row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Summary data
        summary_data = [
            ("BÜYÜKBAŞ", self.report_data["summary"]["buyukbas"]),
            ("KUZU", self.report_data["summary"]["kuzu"]),
            ("OĞLAK", self.report_data["summary"]["oglak"]),
            ("KOYUN", self.report_data["summary"]["koyun"]),
            ("KEÇİ", self.report_data["summary"]["keci"]),
        ]

        summary_row = summary_start_row + 2
        for label, data in summary_data:
            ws.cell(row=summary_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=summary_row, column=2, value=float(data["kesim"]))
            ws.cell(row=summary_row, column=3, value=float(data["deri"]))

            ws.cell(row=summary_row, column=4, value=float(data["bagirsak"]))

            # Apply borders and styling
            for col in range(1, 5):
                cell = ws.cell(row=summary_row, column=col)
                cell.border = thin_border
                if col == 1:
                    cell.font = Font(bold=True)

            summary_row += 1

        # Auto-adjust column widths with a larger cap and explicit width for ID column.
        # Cell.value can be None or non-string; catch TypeError/AttributeError for column-width logic only.
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Column 5 is HAYVAN KİMLİK NO explicitly widen for long tags (moved due to removed column)
        try:
            ws.column_dimensions[get_column_letter(5)].width = max(ws.column_dimensions[get_column_letter(5)].width, 25)
        except (TypeError, AttributeError, KeyError):
            ws.column_dimensions[get_column_letter(5)].width = 25

        self._add_operational_kpis_sheet(wb)
        return wb

    def _add_operational_kpis_sheet(self, wb):
        """Second sheet: pipeline counts, yield, and rollups (mirrors get_all_data extras)."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet(title="Operational KPIs")
        rd = self.report_data
        thin_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

        ws["A1"] = "Operational summary"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Period: {rd.get('start_date', '')} — {rd.get('end_date', '')}"
        row = 4

        def section_title(text):
            nonlocal row
            ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=12)
            row += 1

        def two_col_table(headers, rows_iter):
            nonlocal row
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = Font(bold=True)
                cell.fill = thin_fill
            row += 1
            for r in rows_iter:
                for c, val in enumerate(r, 1):
                    ws.cell(row=row, column=c, value=val)
                row += 1
            row += 1

        status_sd = rd.get("status_breakdown_slaughter_period") or {}
        section_title("Status counts (slaughter date in period)")
        if status_sd:
            two_col_table(["Status", "Headcount"], sorted(status_sd.items(), key=lambda x: x[0]))
        else:
            ws.cell(row=row, column=1, value="No animals with slaughter date in this period.")
            row += 2

        wip = rd.get("wip_intake_by_received_date") or {}
        section_title("WIP — received date in period (received / slaughtered only)")
        if wip:
            two_col_table(["Status", "Headcount"], sorted(wip.items(), key=lambda x: x[0]))
        else:
            ws.cell(row=row, column=1, value="None")
            row += 2

        ys = rd.get("yield_stats") or {}
        section_title("Yield (hot / live × 100, carcass-ready+ rows with both weights)")
        two_col_table(
            ["Metric", "Value"],
            [
                ("Samples", ys.get("sample_count", 0)),
                ("Mean yield %", ys.get("mean_yield_pct") if ys.get("mean_yield_pct") is not None else "—"),
                ("Median yield %", ys.get("median_yield_pct") if ys.get("median_yield_pct") is not None else "—"),
            ],
        )

        by_client = rd.get("by_client") or {}
        section_title("By client (firm)")
        if by_client:
            two_col_table(
                ["Client", "Headcount", "Hot carcass kg"],
                [(k, v["headcount"], round(v["hot_carcass_kg"], 2)) for k, v in by_client.items()],
            )
        else:
            ws.cell(row=row, column=1, value="—")
            row += 2

        by_type = rd.get("by_animal_type") or {}
        section_title("By animal type (report rows)")
        if by_type:
            two_col_table(
                ["Type", "Headcount", "Hot carcass kg"],
                [(k, v["headcount"], round(v["hot_carcass_kg"], 2)) for k, v in by_type.items()],
            )

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 45)

    def generate_cold_shrinkage_excel(self):
        """Generate Excel for cold shrinkage (fire) analysis."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Fire Detay"

        rd = self.report_data
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

        ws["A1"] = f"FIRE ANALİZİ — {rd.get('start_date', '')} / {rd.get('end_date', '')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:G1")

        for c, h in enumerate(
            ["HAYVAN KİMLİK NO", "CİNSİ", "FİRMA", "SICAK KARKAS (kg)", "SOĞUK KARKAS (kg)", "FIRE (kg)", "FIRE (%)"],
            1,
        ):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        data = rd.get("cold_shrinkage", {})
        row = 4
        for i, item in enumerate(data.get("rows", [])):
            fill = alt_fill if i % 2 == 0 else None
            for c, val in enumerate(
                [
                    item.get("identification_tag", ""),
                    item.get("animal_type", ""),
                    item.get("client_name", ""),
                    item.get("hot_carcass_weight", 0),
                    item.get("cold_carcass_weight", 0),
                    item.get("shrinkage_kg", 0),
                    item.get("shrinkage_pct"),
                ],
                1,
            ):
                cell = ws.cell(row=row, column=c, value=val)
                if fill:
                    cell.fill = fill
            row += 1

        row += 1
        for label, val in [
            ("Örnek Sayısı", data.get("sample_count", 0)),
            ("Ort. Fire %", f"{data['mean_shrinkage_pct']:.2f}%" if data.get("mean_shrinkage_pct") is not None else "—"),
            ("Medyan Fire %", f"{data['median_shrinkage_pct']:.2f}%" if data.get("median_shrinkage_pct") is not None else "—"),
        ]:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=val)
            row += 1

        ws2 = wb.create_sheet(title="Özet - Hayvan Tipi")
        ws2["A1"] = "Hayvan Tipine Göre Fire Özeti"
        ws2["A1"].font = Font(bold=True, size=12)
        for c, h in enumerate(["CİNSİ", "ÖRNEK SAYISI", "ORT. FIRE %", "MEDYAN FIRE %"], 1):
            cell = ws2.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
        r2 = 4
        for t, stats in (data.get("by_animal_type") or {}).items():
            ws2.cell(row=r2, column=1, value=t)
            ws2.cell(row=r2, column=2, value=stats.get("sample_count", 0))
            ws2.cell(row=r2, column=3, value=f"{stats['mean_pct']:.2f}%" if stats.get("mean_pct") is not None else "—")
            ws2.cell(row=r2, column=4, value=f"{stats['median_pct']:.2f}%" if stats.get("median_pct") is not None else "—")
            r2 += 1

        for sheet in [ws, ws2]:
            for column in sheet.columns:
                max_length = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                sheet.column_dimensions[col_letter].width = min(max_length + 2, 40)

        return wb

    def generate_cut_yield_excel(self):
        """Generate Excel for disassembly cut yield analysis."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Parça Et Verimi"

        rd = self.report_data
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        alt_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")

        cut_data = rd.get("cut_yield", {})
        total_hot = cut_data.get("total_hot_carcass_kg", 0)

        ws["A1"] = f"PARÇA ET VERİMİ — {rd.get('start_date', '')} / {rd.get('end_date', '')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:E1")
        ws.cell(row=3, column=1, value=f"Toplam Sıcak Karkas: {total_hot:.2f} kg").font = Font(italic=True)
        ws.merge_cells("A3:E3")

        for c, h in enumerate(
            ["PARÇA ADI", "TOPLAM (kg)", "ADET", "SICAK KARKAS (%)", "SCALE SYNCED ADET"],
            1,
        ):
            cell = ws.cell(row=5, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row = 6
        for i, cut in enumerate(cut_data.get("cuts", [])):
            fill = alt_fill if i % 2 == 0 else None
            for c, val in enumerate(
                [
                    cut.get("cut_name", ""),
                    cut.get("total_kg", 0),
                    cut.get("cut_count", 0),
                    f"{cut['pct_of_hot']:.2f}%" if cut.get("pct_of_hot") is not None else "—",
                    cut.get("scale_synced", 0),
                ],
                1,
            ):
                cell = ws.cell(row=row, column=c, value=val)
                if fill:
                    cell.fill = fill
            row += 1

        for column in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        return wb

    def generate_pipeline_time_excel(self):
        """Generate Excel for pipeline time / stage duration analysis."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Aşama Süreleri"

        rd = self.report_data
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7B2C2C", end_color="7B2C2C", fill_type="solid")

        ws["A1"] = f"SÜREÇ SÜRESİ ANALİZİ — {rd.get('start_date', '')} / {rd.get('end_date', '')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        for c, h in enumerate(["AŞAMA", "ÖRNEK SAYISI", "ORT. SÜRE", "MEDYAN SÜRE"], 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        data = rd.get("pipeline_time", {})
        intake = data.get("intake_to_slaughter_hours", {})
        session = data.get("disassembly_session_minutes", {})

        for row, (label, stats, unit) in enumerate(
            [
                ("Kabul → Kesim", intake, "saat"),
                ("Parçalama Süresi", session, "dk"),
            ],
            4,
        ):
            ws.cell(row=row, column=1, value=f"{label} ({unit})")
            ws.cell(row=row, column=2, value=stats.get("sample_count", 0))
            ws.cell(row=row, column=3, value=f"{stats['mean']:.2f}" if stats.get("mean") is not None else "—")
            ws.cell(row=row, column=4, value=f"{stats['median']:.2f}" if stats.get("median") is not None else "—")

        ws2 = wb.create_sheet(title="Durum Dağılımı")
        ws2["A1"] = "Kesim Tarihine Göre Durum Dağılımı"
        ws2["A1"].font = Font(bold=True, size=12)
        for c, h in enumerate(["DURUM", "HAYVAN SAYISI"], 1):
            cell = ws2.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
        r2 = 4
        for status, count in sorted((data.get("animals_by_status") or {}).items()):
            ws2.cell(row=r2, column=1, value=status)
            ws2.cell(row=r2, column=2, value=count)
            r2 += 1

        for sheet in [ws, ws2]:
            for column in sheet.columns:
                max_length = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                sheet.column_dimensions[col_letter].width = min(max_length + 2, 40)

        return wb

    def generate_byproduct_offal_excel(self):
        """Generate Excel for byproduct and offal breakdown."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Sakatat"

        rd = self.report_data
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        data = rd.get("byproduct_offal", {})

        ws["A1"] = f"SAKATAT & YAN ÜRÜNLER — {rd.get('start_date', '')} / {rd.get('end_date', '')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        for c, h in enumerate(["SAKATAT TİPİ", "DURUM", "TOPLAM AĞIRLIK (kg)", "ADET"], 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row = 4
        for item in data.get("offal_by_type", []):
            ws.cell(row=row, column=1, value=item.get("offal_type", ""))
            ws.cell(row=row, column=2, value=item.get("disposition", ""))
            ws.cell(row=row, column=3, value=item.get("total_weight", 0))
            ws.cell(row=row, column=4, value=item.get("count", 0))
            row += 1

        ws2 = wb.create_sheet(title="Yan Ürünler")
        ws2["A1"] = f"YAN ÜRÜNLER — {rd.get('start_date', '')} / {rd.get('end_date', '')}"
        ws2["A1"].font = Font(bold=True, size=14)
        ws2.merge_cells("A1:D1")

        for c, h in enumerate(["YAN ÜRÜN TİPİ", "DURUM", "TOPLAM AĞIRLIK (kg)", "ADET"], 1):
            cell = ws2.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        r2 = 4
        for item in data.get("byproducts_by_type", []):
            ws2.cell(row=r2, column=1, value=item.get("byproduct_type", ""))
            ws2.cell(row=r2, column=2, value=item.get("disposition", ""))
            ws2.cell(row=r2, column=3, value=item.get("total_weight", 0))
            ws2.cell(row=r2, column=4, value=item.get("count", 0))
            r2 += 1

        for sheet in [ws, ws2]:
            for column in sheet.columns:
                max_length = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                sheet.column_dimensions[col_letter].width = min(max_length + 2, 40)

        return wb

    def generate_client_activity_excel(self):
        """Generate Excel for client activity summary."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Müşteri Özeti"

        rd = self.report_data
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        alt_fill = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")

        ws["A1"] = f"MÜŞTERİ ÖZETİ — {rd.get('start_date', '')} / {rd.get('end_date', '')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:E1")

        for c, h in enumerate(
            ["FİRMA / MÜŞTERİ", "DURUM", "SİPARİŞ SAYISI", "HAYVAN SAYISI", "SICAK KARKAS (kg)"],
            1,
        ):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row = 4
        for i, item in enumerate(rd.get("client_activity", [])):
            fill = alt_fill if i % 2 == 0 else None
            for c, val in enumerate(
                [
                    item.get("client_name", ""),
                    item.get("status", ""),
                    item.get("order_count", 0),
                    item.get("animal_count", 0),
                    item.get("total_hot_carcass_kg", 0),
                ],
                1,
            ):
                cell = ws.cell(row=row, column=c, value=val)
                if fill:
                    cell.fill = fill
            row += 1

        for column in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        return wb

    def generate_client_order_receipt_excel(self):
        """Generate Excel for customer portal order receipt."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        first_sheet = True
        rd = self.report_data
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for order in rd.get("client_orders", []):
            order_no = order.get("order_no", "Sipariş")
            if first_sheet:
                ws = wb.active
                ws.title = f"Sipariş {order_no}"[:31]
                first_sheet = False
            else:
                ws = wb.create_sheet(title=f"Sipariş {order_no}"[:31])

            ws["A1"] = f"KESİM RAPORU — Sipariş No: {order_no}"
            ws["A1"].font = Font(bold=True, size=14)
            ws.merge_cells("A1:F1")
            ws["A2"] = (
                f"Tarih: {order.get('order_date', '')}  |  "
                f"Durum: {order.get('status', '')}  |  "
                f"Hizmet: {order.get('service_package', '')}"
            )
            ws.merge_cells("A2:F2")

            for c, h in enumerate(
                ["HAYVAN KİMLİK NO", "CİNSİ", "DURUM", "CANLI AĞIRLIK (kg)", "SICAK KARKAS (kg)", "DERİ (kg)"],
                1,
            ):
                cell = ws.cell(row=4, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            row = 5
            for animal in order.get("animals", []):
                for c, val in enumerate(
                    [
                        animal.get("identification_tag", ""),
                        animal.get("animal_type", ""),
                        animal.get("status", ""),
                        animal.get("live_weight", 0),
                        animal.get("hot_carcass_weight", 0),
                        animal.get("leather_weight", 0),
                    ],
                    1,
                ):
                    ws.cell(row=row, column=c, value=val)
                row += 1
                cuts = animal.get("cuts", [])
                if cuts:
                    row += 1
                    ws.cell(row=row, column=2, value="Parça").font = Font(italic=True, bold=True)
                    ws.cell(row=row, column=3, value="Ağırlık (kg)").font = Font(italic=True, bold=True)
                    row += 1
                    for cut in cuts:
                        ws.cell(row=row, column=2, value=cut.get("cut_name", ""))
                        ws.cell(row=row, column=3, value=cut.get("weight_kg", 0))
                        row += 1
                    row += 1

            for column in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        if first_sheet:
            ws = wb.active
            ws.title = "Kesim Raporu"
            ws["A1"] = "Seçilen tarih aralığında sipariş bulunamadı."
            ws["A1"].font = Font(italic=True)

        return wb


class PDFReportGenerator:
    """Service for generating PDF reports using ReportLab with improved formatting"""

    def __init__(self, report_data):
        self.report_data = report_data

    def _convert_turkish_chars(self, text):
        """Convert Turkish characters to ASCII equivalents"""
        if not text:
            return text

        turkish_to_ascii = {
            "Ğ": "G",
            "ğ": "g",
            "Ü": "U",
            "ü": "u",
            "Ş": "S",
            "ş": "s",
            "İ": "I",
            "ı": "i",
            "Ö": "O",
            "ö": "o",
            "Ç": "C",
            "ç": "c",
        }

        for turkish, ascii_char in turkish_to_ascii.items():
            text = text.replace(turkish, ascii_char)

        return text

    def _truncate_text(self, text, max_length=20):
        """Truncate text to fit in table cells"""
        if not text:
            return ""
        text = str(text)
        if len(text) <= max_length:
            return text
        return text[: max_length - 3] + "..."

    def _wrap_text(self, text, max_chars_per_line=15):
        """Wrap long text into multiple lines"""
        if not text:
            return ""
        text = str(text)
        if len(text) <= max_chars_per_line:
            return text

        words = text.split()
        lines = []
        current_line = ""

        for word in words:
            if len(current_line + " " + word) <= max_chars_per_line:
                if current_line:
                    current_line += " " + word
                else:
                    current_line = word
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
                if len(current_line) > max_chars_per_line:
                    # If single word is too long, truncate it
                    current_line = current_line[: max_chars_per_line - 3] + "..."

        if current_line:
            lines.append(current_line)

        return "\n".join(lines)

    def _build_pdf_kpi_rows(self):
        """Rows for the operational KPI block (mirrors Excel Operational KPIs sheet)."""
        rd = self.report_data
        rows = [["Metric", "Value"]]
        rows.append(["Period", f"{rd.get('start_date', '')} - {rd.get('end_date', '')}"])
        sb = rd.get("status_breakdown_slaughter_period") or {}
        for k, v in sorted(sb.items()):
            rows.append([f"Slaughter period / status={k}", str(v)])
        wip = rd.get("wip_intake_by_received_date") or {}
        for k, v in sorted(wip.items()):
            rows.append([f"WIP intake / status={k}", str(v)])
        ys = rd.get("yield_stats") or {}
        rows.append(["Yield sample count", str(ys.get("sample_count", 0))])
        mean_y = ys.get("mean_yield_pct")
        med_y = ys.get("median_yield_pct")
        if mean_y is not None:
            rows.append(["Mean yield %", f"{mean_y:.2f}"])
        if med_y is not None:
            rows.append(["Median yield %", f"{med_y:.2f}"])
        return rows

    def generate_daily_slaughter_pdf(self):
        """Generate improved PDF for daily slaughter reports with better formatting"""
        import tempfile

        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        temp_file.close()

        # Use landscape orientation for better table layout
        doc = SimpleDocTemplate(
            temp_file.name,
            pagesize=landscape(A4),
            rightMargin=1 * cm,
            leftMargin=1 * cm,
            topMargin=2 * cm,
            bottomMargin=1 * cm,
        )
        story = []

        # Get styles
        styles = getSampleStyleSheet()

        # Company header style
        company_style = ParagraphStyle(
            "CompanyHeader",
            parent=styles["Heading1"],
            fontSize=14,
            spaceAfter=10,
            alignment=TA_CENTER,
            textColor=colors.darkblue,
            fontName="Helvetica-Bold",
        )

        # Title style
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=15,
            alignment=TA_CENTER,
            textColor=colors.darkred,
            fontName="Helvetica-Bold",
        )

        # Date style
        date_style = ParagraphStyle(
            "DateStyle",
            parent=styles["Normal"],
            fontSize=10,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.black,
            fontName="Helvetica",
        )

        # Summary title style
        summary_style = ParagraphStyle(
            "SummaryTitle",
            parent=styles["Heading2"],
            fontSize=12,
            spaceAfter=10,
            alignment=TA_LEFT,
            textColor=colors.darkblue,
            fontName="Helvetica-Bold",
        )

        # Add company header
        company_info = "GUNDOGDULAR GIDA SAN VE TUR. TIC. LTD STI - BOZALAN - EZINE / CANAKKALE"
        company_para = Paragraph(company_info, company_style)
        story.append(company_para)
        story.append(Spacer(1, 10))

        # Add title
        title = Paragraph("GUNLUK KESIM RAPORU", title_style)
        story.append(title)

        # Add date range
        start_date = self.report_data.get("start_date", "")
        end_date = self.report_data.get("end_date", "")
        current_date = datetime.now().strftime("%d.%m.%Y %H:%M")
        date_text = f"Tarih Araligi: {start_date} - {end_date} | Rapor Tarihi: {current_date}"
        date_para = Paragraph(date_text, date_style)
        story.append(date_para)
        story.append(Spacer(1, 15))

        # KPI summary (same metrics as Excel second sheet)
        kpi_rows = self._build_pdf_kpi_rows()
        if len(kpi_rows) > 1:
            kpi_table = Table(kpi_rows, colWidths=[6 * cm, 10 * cm])
            kpi_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            story.append(Paragraph("OPERASYONEL OZET", summary_style))
            story.append(Spacer(1, 6))
            story.append(kpi_table)
            story.append(Spacer(1, 15))

        # Main data table with improved layout
        daily_data = self.report_data.get("daily_data", [])
        if daily_data:
            # Table headers with better column names
            headers = [
                "ALINAN MUSTERI",
                "ADET",
                "CINSI",
                "SICAK KARKAS",
                "HAYVAN KIMLIK NO",
                "SAKATAT",
                "BAGIRSAK",
                "DERI",
                "FIRMA UNVANI",
                "ACIKLAMA",
            ]

            # Prepare table data with text wrapping
            table_data = [headers]
            for item in daily_data:
                # Wrap long text for better display
                destination = self._wrap_text(self._convert_turkish_chars(item.get("destination", "")), 12)
                client_name = self._wrap_text(self._convert_turkish_chars(item.get("client_name", "")), 12)
                description = self._wrap_text(self._convert_turkish_chars(item.get("description", "")), 10)

                row = [
                    destination,
                    str(int(float(item.get("quantity", 0)))),
                    self._convert_turkish_chars(item.get("animal_type", "")),
                    f"{float(item.get('hot_carcass_weight', 0)):.1f}",
                    self._convert_turkish_chars(item.get("identification_tag", "")),
                    self._convert_turkish_chars(item.get("offal_status", "")),
                    self._convert_turkish_chars(item.get("bowels_status", "")),
                    f"{float(item.get('leather_weight', 0)):.1f}",
                    client_name,
                    description,
                ]
                table_data.append(row)

            # Create table with better column widths for Turkish headers
            # Increased identification tag (HAYVAN KIMLIK NO) column width for long values
            col_widths = [4 * cm, 1.5 * cm, 1.5 * cm, 2.5 * cm, 4 * cm, 2 * cm, 2 * cm, 1.5 * cm, 4 * cm, 2.5 * cm]
            table = Table(table_data, colWidths=col_widths, repeatRows=1)

            # Enhanced table styling
            table.setStyle(
                TableStyle(
                    [
                        # Header styling
                        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 0), (-1, 0), 8),
                        ("LEFTPADDING", (0, 0), (-1, 0), 4),
                        ("RIGHTPADDING", (0, 0), (-1, 0), 4),
                        # Data styling
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("TOPPADDING", (0, 1), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
                        ("LEFTPADDING", (0, 1), (-1, -1), 3),
                        ("RIGHTPADDING", (0, 1), (-1, -1), 3),
                        # Alternating row colors for better readability
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.lightgrey, colors.white]),
                        # Special alignment for numeric columns
                        ("ALIGN", (1, 1), (1, -1), "CENTER"),  # ADET
                        ("ALIGN", (3, 1), (3, -1), "CENTER"),  # Hot carcass weight column
                        ("ALIGN", (7, 1), (7, -1), "CENTER"),  # DERI
                    ]
                )
            )

            story.append(table)
            story.append(Spacer(1, 20))

        # Summary section with improved layout
        summary = self.report_data.get("summary", {})
        if summary:
            summary_title = Paragraph("OZET TABLOSU", summary_style)
            story.append(summary_title)

            # Summary table with totals
            summary_headers = ["HAYVAN TURU", "KESIM ADEDI", "DERI (KG)", "BAGIRSAK", "SAKATAT"]
            summary_data = [summary_headers]

            # Calculate totals
            total_kesim = 0
            total_deri = 0
            total_bagirsak = 0
            total_sakatat = 0

            # Add summary rows
            for animal_type, data in summary.items():
                if animal_type in ["buyukbas", "kuzu", "oglak", "koyun", "keci"]:
                    turkish_name = {
                        "buyukbas": "BUYUKBAS",
                        "kuzu": "KUZU",
                        "oglak": "OGLAK",
                        "koyun": "KOYUN",
                        "keci": "KECI",
                    }.get(animal_type, animal_type.upper())

                    kesim = float(data.get("kesim", 0))
                    deri = float(data.get("deri", 0))
                    bagirsak = float(data.get("bagirsak", 0))
                    sakatat = float(data.get("sakatat", 0))

                    total_kesim += kesim
                    total_deri += deri
                    total_bagirsak += bagirsak
                    total_sakatat += sakatat

                    row = [
                        turkish_name,
                        f"{kesim:.0f}",
                        f"{deri:.1f}",
                        f"{bagirsak:.0f}" if animal_type != "keci" else "-",
                        f"{sakatat:.0f}",
                    ]
                    summary_data.append(row)

            # Add totals row
            summary_data.append(
                ["TOPLAM", f"{total_kesim:.0f}", f"{total_deri:.1f}", f"{total_bagirsak:.0f}", f"{total_sakatat:.0f}"]
            )

            # Create summary table
            summary_col_widths = [3 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm]
            summary_table = Table(summary_data, colWidths=summary_col_widths, repeatRows=1)
            summary_table.setStyle(
                TableStyle(
                    [
                        # Header styling
                        ("BACKGROUND", (0, 0), (-1, 0), colors.darkgreen),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 0), (-1, 0), 8),
                        # Data styling
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        # Special styling for totals row
                        ("BACKGROUND", (0, -1), (-1, -1), colors.lightblue),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, -1), (-1, -1), 9),
                        # Alternating row colors
                        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.beige, colors.white]),
                    ]
                )
            )

            story.append(summary_table)

        # Add footer with generation info
        footer_style = ParagraphStyle(
            "FooterStyle",
            parent=styles["Normal"],
            fontSize=8,
            spaceAfter=10,
            alignment=TA_CENTER,
            textColor=colors.grey,
            fontName="Helvetica",
        )
        footer_text = f"Bu rapor {current_date} tarihinde otomatik olarak olusturulmustur."
        footer = Paragraph(footer_text, footer_style)
        story.append(Spacer(1, 20))
        story.append(footer)

        # Build PDF
        doc.build(story)

        return temp_file.name

    def generate_client_order_receipt_pdf(self):
        """Generate PDF receipt for a customer — one section per order."""
        import tempfile

        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        temp_file.close()

        doc = SimpleDocTemplate(
            temp_file.name,
            pagesize=A4,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Title"],
            fontSize=14,
            fontName="Helvetica-Bold",
            spaceAfter=6,
            alignment=TA_CENTER,
        )
        section_style = ParagraphStyle(
            "Section",
            parent=styles["Normal"],
            fontSize=10,
            fontName="Helvetica-Bold",
            spaceAfter=4,
            spaceBefore=10,
        )
        normal_style = ParagraphStyle(
            "Normal8",
            parent=styles["Normal"],
            fontSize=8,
            fontName="Helvetica",
        )

        story = []

        rd = self.report_data
        current_date = self._convert_turkish_chars(
            __import__("django.utils.timezone", fromlist=["timezone"]).now().strftime("%Y-%m-%d %H:%M")
        )

        story.append(Paragraph("KESIM RAPORU", title_style))
        story.append(Spacer(1, 6))

        orders = rd.get("client_orders", [])
        if not orders:
            story.append(Paragraph("Secilen tarih araliginda siparis bulunamadi.", normal_style))
        else:
            for order in orders:
                order_no = self._convert_turkish_chars(order.get("order_no", ""))
                order_date = order.get("order_date", "")
                status = self._convert_turkish_chars(order.get("status", ""))
                service = self._convert_turkish_chars(order.get("service_package", ""))

                story.append(
                    Paragraph(
                        f"Siparis No: {order_no}  |  Tarih: {order_date}  |  Durum: {status}  |  Hizmet: {service}",
                        section_style,
                    )
                )

                animals = order.get("animals", [])
                if animals:
                    animal_headers = [
                        "KIMLIK NO", "CINSI", "DURUM",
                        "CANLI (kg)", "SICAK KARKAS (kg)", "DERI (kg)",
                    ]
                    animal_rows = [animal_headers]
                    for a in animals:
                        animal_rows.append([
                            self._convert_turkish_chars(a.get("identification_tag", "")),
                            self._convert_turkish_chars(a.get("animal_type", "")),
                            self._convert_turkish_chars(a.get("status", "")),
                            f"{float(a.get('live_weight', 0)):.1f}",
                            f"{float(a.get('hot_carcass_weight', 0)):.1f}",
                            f"{float(a.get('leather_weight', 0)):.1f}",
                        ])

                    col_widths = [3 * cm, 2 * cm, 3 * cm, 2.5 * cm, 3.5 * cm, 2 * cm]
                    animal_table = Table(animal_rows, colWidths=col_widths, repeatRows=1)
                    animal_table.setStyle(
                        TableStyle([
                            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 8),
                            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.lightgrey, colors.white]),
                            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 4),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ])
                    )
                    story.append(animal_table)

                    # Cuts sub-tables (only for animals that have cuts)
                    for a in animals:
                        cuts = a.get("cuts", [])
                        if cuts:
                            tag = self._convert_turkish_chars(a.get("identification_tag", "") or a.get("animal_type", ""))
                            story.append(Spacer(1, 4))
                            story.append(Paragraph(f"Parca Detay — {tag}", normal_style))
                            cut_data = [["PARCA ADI", "AGIRLIK (kg)"]]
                            for cut in cuts:
                                cut_data.append([
                                    self._convert_turkish_chars(cut.get("cut_name", "")),
                                    f"{float(cut.get('weight_kg', 0)):.2f}",
                                ])
                            cut_table = Table(cut_data, colWidths=[6 * cm, 3 * cm])
                            cut_table.setStyle(
                                TableStyle([
                                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#375623")),
                                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                                ])
                            )
                            story.append(cut_table)

                story.append(Spacer(1, 16))

        footer_style = ParagraphStyle(
            "Footer",
            parent=styles["Normal"],
            fontSize=7,
            alignment=TA_CENTER,
            textColor=colors.grey,
            fontName="Helvetica",
        )
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"Bu rapor {current_date} tarihinde olusturulmustur.", footer_style))

        doc.build(story)
        return temp_file.name
